import pandas as pd
import os
import re
import unicodedata
import numpy as np
import datetime
from openpyxl import load_workbook
import calendar

class TextUtils:
    @staticmethod
    def normalize_text(text: str) -> str:
        if not isinstance(text, str):
            text = str(text)
        text = text.lower().strip()
        text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
        text = text.replace("\n", " ").replace("/", " ")
        text = re.sub(r"\s+", " ", text)
        return text

    @staticmethod
    def contains_whole_word(text: str, phrase: str) -> bool:
        text = TextUtils.normalize_text(text)
        phrase = TextUtils.normalize_text(phrase)
        pattern = r'(?<!\w)' + re.escape(phrase) + r'(?!\w)'
        return re.search(pattern, text) is not None


class ColumnMatcher:
    @staticmethod
    def score_match(col: str, keywords, priority=None, avoid=None) -> int:
        if priority is None: priority = []
        if avoid is None: avoid = []

        txt = TextUtils.normalize_text(col)
        score = 0

        for p in priority:
            p_n = TextUtils.normalize_text(p)
            if txt == p_n:
                score = max(score, 100)
            elif TextUtils.contains_whole_word(txt, p_n):
                score = max(score, 90)
            elif p_n in txt:
                score = max(score, 80)

        for k in keywords:
            k_n = TextUtils.normalize_text(k)
            if txt == k_n:
                score = max(score, 70)
            elif TextUtils.contains_whole_word(txt, k_n):
                score = max(score, 60)
            elif k_n in txt:
                score = max(score, 40)

        for a in avoid:
            a_n = TextUtils.normalize_text(a)
            if TextUtils.contains_whole_word(txt, a_n) or a_n in txt:
                score -= 50

        return score

    @staticmethod
    def find_best_match(columns, keywords, priority=None, avoid=None):
        if priority is None: priority = []
        if avoid is None: avoid = []

        scored = [(col, ColumnMatcher.score_match(col, keywords, priority, avoid)) 
                  for col in columns if ColumnMatcher.score_match(col, keywords, priority, avoid) > 0]

        if not scored:
            return None, []

        scored.sort(key=lambda x: (-x[1], len(str(x[0]))))
        return scored[0][0], scored


class DataProcessor:
    # unified hyperlink-bearing headers (normalized)
    HYPERLINK_HEADERS = {
        "poza", "photo", "schita", "link poza", "foto", "link foto", "imagini", "picture",
        "catalog", "photo/video", "photo & map",
        "cod", "pagina prezentare", "code", "poza locatie", "imagine", "imagini 1",
        "link poza suport publicitar", "foto locatie", "adresa", "site"
    }

    TECH_DETAILS_HEADERS = {
        "schita", "foto schita", "production sketch", "schita productie",
        "schita de productie pentru material publicitar",
        "link schita productie", "technical details"
    }

    URL_RE = re.compile(r"^(?:https?://|www\.)", re.IGNORECASE)
    HYPERLINK_FORMULA_RE = re.compile(r'^\s*=\s*HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)

    def __init__(self, groups, directory=None):
        self.groups = groups
        self.directory = directory

    @staticmethod
    def _extract_hyperlinks(file_input, sheet_index=0):
        """
        Returns {(row0, col0): url} for the given sheet.
        Supports both cell.hyperlink and =HYPERLINK("url","text").
        Indices are 0-based to match pandas.
        """
        if hasattr(file_input, "seek"):
            file_input.seek(0)
        # data_only=False so we can read formulas
        wb = load_workbook(file_input, data_only=False, read_only=False)
        # use the same sheet as pandas (index 0)
        ws = wb.worksheets[sheet_index]

        links = {}
        for r, row in enumerate(ws.iter_rows()):
            for c, cell in enumerate(row):
                url = None
                if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                    url = cell.hyperlink.target
                else:
                    v = cell.value
                    if isinstance(v, str):
                        m = DataProcessor.HYPERLINK_FORMULA_RE.match(v)
                        if m:
                            url = m.group(1)
                if url:
                    links[(r, c)] = url
        return links

    def extract_standardized_dataframe(self, file_input, file_name=None):
        """
        Reads Excel/CSV, matches columns using groups.json, and returns a standardized DataFrame.
        Consolidates all relevant hyperlinks into "Photo Link" and "Tech Details" columns.
        """
        if isinstance(file_input, (str, bytes, os.PathLike)):
            file_name = os.path.basename(file_input)
            read_target = file_input
        else:
            read_target = file_input
            if file_name is None:
                file_name = "uploaded_file.xlsx"

        # 1) sniff header row
        try:
            if hasattr(read_target, "seek"):
                read_target.seek(0)
            raw_data = pd.read_excel(read_target, sheet_name=0, header=None, engine="openpyxl")
        except Exception as e:
            print(f"Failed to read {file_name}: {e}")
            return pd.DataFrame()

        header_row_index = None
        for i, row in raw_data.iterrows():
            if row.count() >= 9:
                header_row_index = i
                break
        if header_row_index is None:
            return pd.DataFrame()

        # 2) read full sheet with headers
        if hasattr(read_target, "seek"):
            read_target.seek(0)
        df = pd.read_excel(read_target, sheet_name=0, header=header_row_index, engine="openpyxl")
        columns = df.columns

        # 3) pull hyperlinks from the SAME sheet (sheet 0), with formula support
        if hasattr(read_target, "seek"):
            read_target.seek(0)
        hyperlinks = self._extract_hyperlinks(read_target, sheet_index=0)

        # 4) build extracted output
        extracted = pd.DataFrame()
        extracted["Photo Link"] = ""   # canonical column for URLs only
        extracted["Tech Details"] = "" # canonical column for tech-related URLs

        # 4a) copy non-Photo-Link / non-Tech-Details groups
        for name, cfg in self.groups.items():
            normalized_name = TextUtils.normalize_text(name)
            if normalized_name in ("photo link", "tech details"):
                continue

            best, _ = ColumnMatcher.find_best_match(
                columns,
                keywords=cfg.get("keywords", []),
                priority=cfg.get("priority", []),
                avoid=cfg.get("avoid", [])
            )
            if best:
                extracted[name] = df[best]
            else:
                extracted[name] = ""

        # 4b) candidate columns for hyperlinks
        candidate_cols_photo = [col for col in df.columns if TextUtils.normalize_text(col) in self.HYPERLINK_HEADERS]

        TECH_DETAILS_HEADERS = {
            "schita", "foto schita", "production sketch", "schita productie",
            "schita de productie pentru material publicitar",
            "link schita productie", "technical details"
        }
        candidate_cols_tech = [col for col in df.columns if TextUtils.normalize_text(col) in TECH_DETAILS_HEADERS]

        candidate_col_indices_photo = [df.columns.get_loc(c) for c in candidate_cols_photo]
        candidate_col_indices_tech  = [df.columns.get_loc(c) for c in candidate_cols_tech]

        # 4c) fill Photo Link and Tech Details
        for row_idx in range(len(df)):
            xl_row0 = header_row_index + 1 + row_idx
            # --- Photo Link ---
            found_photo = []
            for col_idx in candidate_col_indices_photo:
                url = hyperlinks.get((xl_row0, col_idx))
                if url:
                    found_photo.append(url)
            if not found_photo:
                for col_idx in candidate_col_indices_photo:
                    val = df.iloc[row_idx, col_idx]
                    if isinstance(val, str) and self.URL_RE.match(val.strip()):
                        url = val.strip()
                        if url.lower().startswith("www."):
                            url = "http://" + url
                        found_photo.append(url)
            if found_photo:
                extracted.at[row_idx, "Photo Link"] = found_photo[0]

            # --- Tech Details ---
            found_tech = []
            for col_idx in candidate_col_indices_tech:
                url = hyperlinks.get((xl_row0, col_idx))
                if url:
                    found_tech.append(url)
            if not found_tech:
                for col_idx in candidate_col_indices_tech:
                    val = df.iloc[row_idx, col_idx]
                    if isinstance(val, str) and self.URL_RE.match(val.strip()):
                        url = val.strip()
                        if url.lower().startswith("www."):
                            url = "http://" + url
                        found_tech.append(url)
            if found_tech:
                extracted.at[row_idx, "Tech Details"] = found_tech[0]

        # 5) track source file
        extracted["__source_file"] = file_name
        return extracted


    @staticmethod
    def remove_empty_columns(df, excepted_columns):
        cols_to_check = [col for col in df.columns if col != excepted_columns]
        df = df.replace(r'^\s*$', np.nan, regex=True).infer_objects(copy=False)
        return df.dropna(how="all", subset=cols_to_check)

    @staticmethod
    def normalize_dimension(value: str):
        if pd.isna(value) or str(value).strip() == "":
            return None
        value = str(value).strip().lower()
        value = value.replace(",", ".")
        value = re.sub(r"\s*m", "", value)
        return value

    @staticmethod
    def process_size_base_height(row):
        base = DataProcessor.normalize_dimension(row.get("Base"))
        height = DataProcessor.normalize_dimension(row.get("Height"))
        size = str(row.get("Size")) if not pd.isna(row.get("Size")) else ""

        if size.strip():
            parts = re.split(r"[xX]", size)
            if len(parts) == 2:
                base, height = DataProcessor.normalize_dimension(parts[0]), DataProcessor.normalize_dimension(parts[1])
        elif base and height:
            size = f"{base}m x {height}m"

        return pd.Series({
            "Base": f"{base}m" if base else None,
            "Height": f"{height}m" if height else None,
            "Size": size if size else None
        })

    def process_dates(self, df):
        df["Start"] = df["Start"].apply(self._safe_to_date)
        df["End"]   = df["End"].apply(self._safe_to_date)
        return df

    @staticmethod
    def check_if_literal_date(value):
        pattern = r'^\d{1,2}\s+\w+\s*-\s*\d{1,2}\s+\w+$'
        return bool(re.match(pattern, str(value).strip(), re.IGNORECASE))

    @staticmethod
    def split_literal_date(value):
        dates_splitted = [" ".join((d.strip(), str(datetime.datetime.now().year)))
                          for d in str(value).split('-')]
        if len(dates_splitted) == 2:
            return dates_splitted
        return [None, None]

    def deal_with_literal_dates(self, df):
        """
        Extract the first 'Disponibil' period from multi-line text if Start/End are messy.
        """

        def extract_disponibil(x, current_end):
            if pd.isna(x):
                return pd.Series([None, None])

            x_str = str(x)

            # look for 'Disponibil:' line and capture first date range
            match = re.search(r"Disponibil:\s*(\d{2}/\d{2}/\d{2})\s*:\s*(\d{2}/\d{2}/\d{2})", x_str)
            if match:
                start_str, end_str = match.groups()
                try:
                    start = pd.to_datetime(start_str, dayfirst=True).strftime("%Y-%m-%d")
                    end = pd.to_datetime(end_str, dayfirst=True).strftime("%Y-%m-%d")
                    return pd.Series([start, end])
                except:
                    return pd.Series([x, current_end])
            else:
                return pd.Series([x, current_end])

        df[["Start", "End"]] = df.apply(lambda row: extract_disponibil(row["Start"], row["End"]), axis=1)
        return df

    @staticmethod
    def _safe_to_date(value):
        try:
            dt = pd.to_datetime(value, dayfirst=True, errors="raise")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return value

    @staticmethod
    def calculate_no_of_months(row):
        try:
            start = pd.to_datetime(row["Start"], errors="coerce")
            end = pd.to_datetime(row["End"], errors="coerce")
            if pd.isna(start) or pd.isna(end) or start > end:
                return None

            # First month fraction
            days_in_start_month = calendar.monthrange(start.year, start.month)[1]
            first_month_fraction = (days_in_start_month - start.day + 1) / days_in_start_month

            # Last month fraction
            days_in_end_month = calendar.monthrange(end.year, end.month)[1]
            last_month_fraction = end.day / days_in_end_month

            # Full months in between
            full_months = 0
            # Move to next month after start
            current = pd.Timestamp(year=start.year, month=start.month, day=1) + pd.offsets.MonthBegin(1)
            # Loop until the start of end month
            while current < pd.Timestamp(year=end.year, month=end.month, day=1):
                full_months += 1
                current += pd.offsets.MonthBegin(1)

            total_months = first_month_fraction + full_months + last_month_fraction
            return round(total_months, 2)

        except Exception:
            return None
    
    @staticmethod
    def calculate_area_from_size(size_str):
        if pd.isna(size_str):
            return None

        # normalize: lowercase, unify separators, remove units and spaces
        cleaned = (
            str(size_str)
            .lower()
            .replace(",", ".")
            .replace("×", "x")
            .replace("m", "")
            .replace(" ", "")
        )

        # regex: look for two numbers separated by 'x'
        match = re.search(r"(\d*\.?\d+)[xX](\d*\.?\d+)", cleaned)
        if match:
            try:
                base = float(match.group(1))
                height = float(match.group(2))
                return round(base * height, 2)
            except:
                return None
        return None

    @staticmethod
    def build_gps_from_lat_long(row):
        latitude = row.get("Latitude")
        longitude = row.get("Longitude")
        gps_coordinates = row.get("GPS")

        if pd.notna(gps_coordinates) and str(gps_coordinates).strip():
            return gps_coordinates
        if pd.notna(latitude) and pd.notna(longitude):
            return f"{latitude}, {longitude}"
        return None

    def process_files(self, file_objs):
        all_data = []
        for file_input, file_name in file_objs:
            extracted = self.extract_standardized_dataframe(file_input, file_name=file_name)
            if not extracted.empty:
                all_data.append(extracted)

        if not all_data:
            return pd.DataFrame()

        final_df = pd.concat(all_data, ignore_index=True)
        final_df.dropna(how="all", inplace=True)
        final_df = self.remove_empty_columns(final_df, excepted_columns="__source_file")
        final_df[["Base", "Height", "Size"]] = final_df.apply(self.process_size_base_height, axis=1)
        final_df["GPS"] = final_df.apply(self.build_gps_from_lat_long, axis=1)
        final_df = self.process_dates(final_df)
        final_df = self.deal_with_literal_dates(final_df)
        final_df["No. of months"] = final_df.apply(self.calculate_no_of_months, axis=1)
        final_df["NUME FURNIZOR"] = (
            final_df["__source_file"]
            .astype(str)
            .str.strip()
            .str.replace(r"\.xlsx$", "", regex=True, case=False)
            .str.rstrip("-")
            .str.strip()
        )
        final_df = final_df.drop(columns=["Latitude", "Longitude"], errors="ignore")
        final_df["Format"] = final_df["Size"]
        final_df["Size"] = final_df["Size"].apply(self.calculate_area_from_size)
        return final_df
