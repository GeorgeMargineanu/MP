import streamlit as st
import pandas as pd
import json
import io
import time
from logic import DataProcessor
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Data Processor",
    page_icon="static/images/logo.png",  # favicon
    layout="wide"
)

st.sidebar.image("static/images/logo.png", width=120)  

st.title("MP Data Processor")
st.caption(
    "Upload multiple Excel/CSV files, process them with your `groups.json` config, "
    "and download a clean, standardized dataset."
)

st.sidebar.header("⚙️ Configuration")
groups_file = st.sidebar.file_uploader("Upload groups.json", type="json")

uploaded_files = st.file_uploader(
    "Upload Excel/CSV files",
    type=["xlsx", "csv"],
    accept_multiple_files=True
)

# 🔹 Initialize session_state storage
if "final_df" not in st.session_state:
    st.session_state.final_df = None
if "file_objs" not in st.session_state:
    st.session_state.file_objs = []

@st.cache_data
def _read_csv(uploaded_file):
    return pd.read_csv(uploaded_file)

if groups_file and uploaded_files and st.session_state.final_df is None:
    groups = json.load(groups_file)

    file_objs = []
    for uploaded_file in uploaded_files:
        if uploaded_file.type == "text/csv" or uploaded_file.name.endswith(".csv"):
            df = _read_csv(uploaded_file)
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine="openpyxl")
            buffer.seek(0)
        else:
            buffer = io.BytesIO(uploaded_file.read())
            buffer.seek(0)
        file_objs.append((buffer, uploaded_file.name))

    processor = DataProcessor(groups)

    st.subheader("⚙️ Processing Files")
    progress_bar = st.progress(0)
    status_placeholder = st.empty()

    all_results = []
    for i, (buffer, name) in enumerate(file_objs, start=1):
        status_placeholder.info(f"Processing **{name}** ({i}/{len(file_objs)})...")
        df_chunk = processor.process_files([(buffer, name)])
        if not df_chunk.empty:
            all_results.append(df_chunk)
        progress_bar.progress(int(i / len(file_objs) * 100))
        time.sleep(0.2)

    status_placeholder.success("Processing complete!")

    if all_results:
        st.session_state.final_df = pd.concat(all_results, ignore_index=True)
        st.session_state.file_objs = file_objs
    else:
        st.warning(" No data could be extracted.")

# 🔹 Display if already processed
if st.session_state.final_df is not None:
    final_df = st.session_state.final_df
    file_objs = st.session_state.file_objs

    tab1, tab2, tab3 = st.tabs(["🔍 Preview", "📥 Download", "📈 Summary"])

    with tab1:
        st.sidebar.markdown('-----------------')
        options = final_df["__source_file"].unique()

        select_provider = st.sidebar.selectbox("Select the provider",
                                               options=options)
        
        st.write("🔍 Summary for the selected provider :")
        df_per_company = final_df[final_df["__source_file"] == select_provider]
        st.dataframe(df_per_company, use_container_width=True)

    with tab2:
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Processed Data")

            workbook = writer.book
            worksheet = writer.sheets["Processed Data"]

            # style header
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_num, col_name in enumerate(final_df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            max_col_width = 40
            # auto column width, wrap text, zebra striping
            for col_num, col_name in enumerate(final_df.columns, 1):
                col_letter = get_column_letter(col_num)
                max_length = max(final_df[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[col_letter].width = min(max_length, max_col_width)

                for row_idx, row in enumerate(worksheet[col_letter], start=1):
                    row.alignment = Alignment(wrap_text=True, vertical="top")
                    if row_idx > 1:  # zebra striping, skip header
                        if row_idx % 2 == 0:
                            row.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        output_buffer.seek(0) 

        st.download_button(
            label="Download Excel",
            data=output_buffer,
            file_name="processed_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with tab3:
        st.write("### Quick Summary")
        col1, col2, col3 = st.columns(3)
        col1.metric("📂 Files Processed", len(file_objs))
        col2.metric("📊 Rows Combined", len(final_df))
        col3.metric("🧾 Columns Detected", len(final_df.columns))
else:
    st.info("⬅ Please upload a `groups.json` file and at least one Excel/CSV file to begin.")
