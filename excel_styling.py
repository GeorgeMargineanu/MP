from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
import io

def style_and_export_excel(df: pd.DataFrame, metadata: dict) -> io.BytesIO:
    # --- Desired column order (will keep only those present) ---
    columns_ordered = [
        "County", "City", "Address", "ID", "IDF", "Panel type",
        "Format", "Base", "Height", "Size", "Faces", "Start", "End",
        "No. of months", "Rent/month", "Total rent", "Production",
        "Posting", "Ag Comm %", "Agency commission", "Advertising taxe %",
        "Advertising taxe", "Total Cost", "Photo Link", "GPS", "Sketch name",
        "Tech Details", "idx", "NUME FURNIZOR", "CHIRIE FURNIZOR", "POSTARE FURNIZOR",
        "COST PRODUCTIE", "TIP MATERIAL", "__source_file"
    ]
    df = df[[col for col in columns_ordered if col in df.columns]].copy()

    # --- Guardrails: required columns for formulas ---
    required = {
        "Start", "End", "No. of months", "Rent/month", "Total rent", "Production",
        "Posting", "Ag Comm %", "Agency commission", "Advertising taxe %",
        "Advertising taxe", "Total Cost", "Size", "CHIRIE FURNIZOR", "POSTARE FURNIZOR"
    }
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        # Write DataFrame starting at row 10 (0-indexed startrow=9)
        df.to_excel(writer, index=False, sheet_name="Processed Data", startrow=9)
        workbook = writer.book
        worksheet = writer.sheets["Processed Data"]

        # --- Hide default gridlines ---
        worksheet.sheet_view.showGridLines = False

        # --- Insert Image ---
        image_path = "static/images/for_excel_output.png"
        try:
            img = XLImage(image_path)
            img.anchor = "O2"
            img.height = img.height / 2
            img.width = img.width / 2
            worksheet.add_image(img)
        except Exception as e:
            # Non-fatal
            print(f"Could not insert image: {e}")

        # --- Build header -> column letter map dynamically ---
        HEADER_ROW = 10  # because df headers land on row 10 when startrow=9
        header_to_letter = {}
        for col_idx, cell in enumerate(worksheet[HEADER_ROW], start=1):
            if cell.value:
                header_to_letter[str(cell.value).strip()] = get_column_letter(col_idx)

        # Convenience lookups (raise KeyError if missing -> early and loud)
        col_size           = header_to_letter["Size"]                      # J (expected)
        col_start          = header_to_letter["Start"]                     # L
        col_end            = header_to_letter["End"]                       # M
        col_no_months      = header_to_letter["No. of months"]             # N
        col_rent_month     = header_to_letter["Rent/month"]                # O
        col_total_rent     = header_to_letter["Total rent"]                # P
        col_production     = header_to_letter["Production"]                # Q
        col_posting        = header_to_letter["Posting"]                   # R
        col_ag_comm_pct    = header_to_letter["Ag Comm %"]                 # S
        col_ag_comm        = header_to_letter["Agency commission"]         # T
        col_adv_tax_pct    = header_to_letter["Advertising taxe %"]        # U
        col_adv_tax        = header_to_letter["Advertising taxe"]          # V
        col_total_cost     = header_to_letter["Total Cost"]                # W
        col_chirie_furn    = header_to_letter["CHIRIE FURNIZOR"]           # AD (expected)
        col_postare_furn   = header_to_letter["POSTARE FURNIZOR"]          # AE (expected)

        # --- Apply formulas & date values ---
        start_row = 11
        last_row = start_row + len(df) - 1

        for i, row in enumerate(df.itertuples(index=False), start=start_row):
            # Start date
            start_date = pd.to_datetime(getattr(row, "Start"), errors="coerce")
            if not pd.isna(start_date):
                c = worksheet[f"{col_start}{i}"]
                c.value = start_date
                c.number_format = "d-mmm-yy"

            # End date
            end_date = pd.to_datetime(getattr(row, "End"), errors="coerce")
            if not pd.isna(end_date):
                c = worksheet[f"{col_end}{i}"]
                c.value = end_date
                c.number_format = "d-mmm-yy"

            # Rent/month = CHIRIE FURNIZOR * 1.2
            worksheet[f"{col_rent_month}{i}"].value = f"={col_chirie_furn}{i}*1.2"

            # Production = Size * 5
            worksheet[f"{col_production}{i}"].value = f"={col_size}{i}*5"

            # Posting = POSTARE FURNIZOR * 1.2
            worksheet[f"{col_posting}{i}"].value = f"={col_postare_furn}{i}*1.2"

            # Total rent = Rent/month * No. of months
            worksheet[f"{col_total_rent}{i}"].value = f"={col_rent_month}{i}*{col_no_months}{i}"

            # Agency commission = (Posting + Production + Total rent) * Ag Comm %
            worksheet[f"{col_ag_comm}{i}"].value = (
                f"=({col_posting}{i}+{col_production}{i}+{col_total_rent}{i})*{col_ag_comm_pct}{i}"
            )

            # Advertising taxe = (((Total rent + Posting) * Ag Comm %) + Total rent + Posting) * 0.03
            worksheet[f"{col_adv_tax}{i}"].value = (
                f"=((({col_total_rent}{i}+{col_posting}{i})*{col_ag_comm_pct}{i})+{col_total_rent}{i}+{col_posting}{i})*0.03"
            )

            # Total Cost = Adv. taxe % + Agency commission + Posting + Production + Total rent
            # (Kept exactly as in your original logic)
            worksheet[f"{col_total_cost}{i}"].value = (
                f"={col_adv_tax_pct}{i}+{col_ag_comm}{i}+{col_posting}{i}+{col_production}{i}+{col_total_rent}{i}"
            )

            # No. of months (kept your Excel formula, but with dynamic letters)
            worksheet[f"{col_no_months}{i}"].value = (
                f'=IF(OR({col_start}{i}="", {col_end}{i}="", {col_start}{i}>{col_end}{i}), "", '
                f'ROUND((DAY(EOMONTH({col_start}{i},0))-DAY({col_start}{i})+1)/DAY(EOMONTH({col_start}{i},0)) + '
                f'IF(AND(YEAR({col_start}{i})=YEAR({col_end}{i}), MONTH({col_start}{i})=MONTH({col_end}{i})), 0, '
                f'DATEDIF(EOMONTH({col_start}{i},0)+1, DATE(YEAR({col_end}{i}), MONTH({col_end}{i}), 1), "m")) + '
                f'DAY({col_end}{i})/DAY(EOMONTH({col_end}{i},0)), 2))'
            )

        # --- Number formats (run ONCE, outside the row loop) ---
        # Currency columns:
        currency_cols = [
            "Rent/month", "Total rent", "Production", "Posting",
            "Agency commission", "Advertising taxe", "Total Cost"
        ]
        currency_letters = [header_to_letter[c] for c in currency_cols if c in header_to_letter]

        for col_letter in currency_letters:
            for r in range(start_row, last_row + 1):
                cell = worksheet[f"{col_letter}{r}"]
                if cell.value is not None:
                    cell.number_format = '€#,##0.00'

        # Percentage columns:
        pct_cols = ["Ag Comm %", "Advertising taxe %"]
        pct_letters = [header_to_letter[c] for c in pct_cols if c in header_to_letter]

        for col_letter in pct_letters:
            for r in range(start_row, last_row + 1):
                cell = worksheet[f"{col_letter}{r}"]
                if cell.value is not None:
                    cell.number_format = '0.00%'

        # --- Styles ---
        title_font = Font(name="Calibri", size=9, bold=True, color="000000")
        header_font_white = Font(name="Calibri", size=9, color="FFFFFF", bold=True)
        header_font_red = Font(name="Calibri", size=9, color="FF0000", bold=True)
        body_font = Font(name="Calibri", size=9)
        hyperlink_font = Font(name="Calibri", size=9, color="0000FF", underline="single")

        red_fill = PatternFill(start_color="F05055", end_color="F05055", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        special_columns = {
            "idx", "NUME FURNIZOR", "CHIRIE FURNIZOR",
            "COST PRODUCTIE", "TIP MATERIAL", "POSTARE FURNIZOR", "__source_file"
        }

        max_col_width = 40

        # --- Title Row ---
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        c_title = worksheet.cell(row=1, column=1)
        c_title.value = "MP OOH CAMPAIGN"
        c_title.font = title_font
        c_title.alignment = Alignment(horizontal="center", vertical="center")
        c_title.border = thin_border  # border for A1:B1

        # --- Metadata Rows (A2:B7) ---
        meta_fields = ["Client", "Brand", "Campaign", "Version", "Start", "End"]
        for i, field in enumerate(meta_fields, start=2):
            c1 = worksheet.cell(row=i, column=1, value=field)
            c1.font = Font(bold=True, name="Calibri", size=9)
            c1.border = thin_border
            c2 = worksheet.cell(row=i, column=2, value=metadata.get(field, ""))
            c2.font = body_font
            c2.border = thin_border

        # --- Determine hyperlink columns dynamically ---
        hyperlink_columns = [
            col for col in df.columns
            if any(k in col.lower() for k in ["photo", "link", "address", "tech details"])
        ]

        # --- Format DataFrame area (headers + body) ---
        for col_num, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            for row_idx, cell in enumerate(worksheet[col_letter], start=1):

                # Rows 1–9, columns C→last column: untouched
                if 1 <= row_idx <= 9 and col_num >= 3:
                    continue

                # Header row
                if row_idx == HEADER_ROW:
                    if col_name in special_columns:
                        cell.fill = yellow_fill
                        cell.font = header_font_red
                    else:
                        cell.fill = red_fill
                        cell.font = header_font_white
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    worksheet.row_dimensions[cell.row].height = 35

                # Body rows
                elif row_idx > HEADER_ROW:
                    if col_name in hyperlink_columns and cell.value:
                        link = str(cell.value).strip()
                        if link.lower().startswith(("http://", "https://", "www.")):
                            if "tech details" in col_name.lower():
                                display_text = "sketch"
                            elif "address" in col_name.lower():
                                display_text = "link"
                            else:
                                display_text = "photo"
                            if not link.lower().startswith(("http://", "https://")):
                                link = "http://" + link
                            cell.value = display_text
                            cell.hyperlink = link
                            cell.font = hyperlink_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.font = body_font
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        cell.font = body_font
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                    cell.border = thin_border  # body borders

            # Auto column width
            max_length = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[col_letter].width = min(max_length, max_col_width)

    output_buffer.seek(0)
    return output_buffer
